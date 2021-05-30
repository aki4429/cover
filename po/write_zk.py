#!/usr/bin/env python3
# -*- coding: utf-8 -*-

#在庫報告/発注検討表を作成するプログラム。

ZEXCEL = 'TFC_zaiko.xlsx'
KEXCEL = 'TFC_kento_template2.xlsx'

import openpyxl
import csv
import os
from django.conf import settings

from .get_code import get_k, get_z, ori_sort
from .make_balance import MakeBalance
from .index_tool import get_xindex, get_yindex

from po.models import Shouhi


def join_data(c_data, k_data):
    #同じコードのコードデータに検討表データを繋ぐ(左結合)
    jdata = []
    for c_row in c_data:
        c_row = list(c_row)
        flag = 0
        for k_row in k_data:
            if c_row[0] == k_row[0]:
                c_row = c_row + k_row[1:]
                flag = 1

        if flag == 0:
            #同じコードが無い場合、在庫、受注数は0にしておく。
            c_row = c_row + [0,0]

        jdata.append(c_row)

    return jdata

def make_hyo(nolist, codelist, totallist):
    l = len(codelist[0])
    hyo = [['' for i in range(len(nolist)+l)] for j in range(len(codelist)+3)]
    #print('hyo', hyo)
    #先頭列にコードデータを代入
    for i, code in enumerate(codelist):
        hyo[i+3][:0] = code

    #print('hyo_withcode', hyo)
    #１行目にINV no. PO no. を代入
    for i, num in enumerate(nolist):
        hyo[0][i+l] = num[2]

    #2行目にetd を代入
    for i, num in enumerate(nolist):
        hyo[1][i+l] = num[1]

    #3行目にdelivery を代入
    for i, num in enumerate(nolist):
        hyo[2][i+l] = num[0]

    #print('hyo_final', hyo)
    #with open('hyo_final.csv', 'w', encoding='CP932') as f:
    #    writer = csv.writer(f)
    #    writer.writerows(hyo)

    #print('hyo_final.csv を書き出しました。')

    for row in totallist: #着日, etd, PO No. コード　残数
        yindex = get_yindex(hyo, row[3])
        #print('row[3]=', row[3], row[0], row[1], row[2], row[4], yindex)
        #if yindex == 0:
            #print('yindex0=', row[3], row[0], row[1], row[2], row[4])

        if yindex != None and yindex > 2: #yindex 行が 0-2 はタイトル行なので飛ばす。
            #print('index', yindex, row[2], row[4])
            hyo[yindex][get_xindex(hyo, row[2])] = row[4]

        elif yindex != None and yindex < 3:
            print('yindex_<3', yindex, row[2], row[4])


    return hyo

def write_zexcel(hyo, kijunbi, nolist):
    filename = os.path.join(settings.MEDIA_ROOT, 'template', ZEXCEL)
    wb = openpyxl.load_workbook(filename)
    sheet = wb['zaiko']
    sheet['C1'] = kijunbi

    #1行目にinv no, po no 記入
    n = 0 #7列目からスタート
    while n < len(nolist) :
        sheet.cell(row=1, column=n+7, value = nolist[n][2]) #po no.
        #etd, delivery の日付はdatetime.dateなので、strで変換して、
        #何月何日だけ取り出す。
        sheet.cell(row=2, column=n+7, value = str(nolist[n][1])[5:]) #ETD
        sheet.cell(row=3, column=n+7, value = str(nolist[n][0])[5:]) #delivery
        n += 1
    

    #コード他記入
    j=0
    i=4 #4行目からスタート
    while i < len(hyo):
        sheet.cell(row=i, column=2, value = hyo[i-1][0]) #コード
        sheet.cell(row=i, column=3, value = hyo[i-1][2]) #在庫
        sheet.cell(row=i, column=4, value = hyo[i-1][3]) #受注
        sheet.cell(row=i, column=5, value = '=C{0}-D{0}'.format(i)) #有効残
        sheet.cell(row=i, column=6, value = hyo[i-1][1]) #カテゴリー
        #入荷予定記入
        k = 4
        while k < len(hyo[0]): 
            sheet.cell(row=i, column=k+3, value = hyo[i-1][k])
            k+=1
        i += 1
        j += 1

    save_file = 'TFC_zaiko_{0}.xlsx'.format(str(kijunbi).replace('/', '-'))
    #wb.save(save_file)
    #print("{}を保存しました。".format(save_file))

    return wb, save_file

def write_kexcel(hyo, kijunbi, nolist):
    filename = os.path.join(settings.MEDIA_ROOT, 'template', KEXCEL)
    wb = openpyxl.load_workbook(filename)
    sheet = wb['kento']
    sheet['D2'] = kijunbi

    #2行目にinv no, po no 記入
    n = 0 #9列目からスタート
    while n < len(nolist) :
        sheet.cell(row=2, column=n+11, value = nolist[n][2]) #po no.
        sheet.cell(row=3, column=n+11, value = str(nolist[n][1])[5:]) #ETD
        sheet.cell(row=4, column=n+11, value = str(nolist[n][0])[5:]) #delivery
        n += 1
    

    #コード他記入
    j=0
    i=4 #5行目からスタート
    while i < len(hyo):
        sheet.cell(row=i+1, column=2, value = hyo[i-1][0]) #コード
        sheet.cell(row=i+1, column=5, value = hyo[i-1][3]) #在庫
        sheet.cell(row=i+1, column=6, value = hyo[i-1][4]) #受注
        sheet.cell(row=i+1, column=7, value = '=E{0}-F{0}'.format(i+1)) #有効残
        sheet.cell(row=i+1, column=1, value = hyo[i-1][2]) #カテゴリー
        sheet.cell(row=i+1, column=4, value = hyo[i-1][1]) #容積
        #入荷予定記入
        k = 5
        while k < len(hyo[0]): 
            sheet.cell(row=i+1, column=k+6, value = hyo[i-1][k])
            k+=1
        i += 1
        j += 1

    sheet2 = wb['jisseki']
    #コードリスト取り出し
    #検討表書き込み用データhyoからコードだけの
    #リスト(codelist)を取り出す。
    codelist = []
    for i, row in enumerate(hyo):
        if i >2 :
            codelist.append(row[0])

    #重複しない会計年月取得、こんな感じ=>[('1912',), ('2001',), ('2002',)]
    month_list = Shouhi.objects.all().order_by('month').distinct().values_list('month')

    #消費実績表作成 
    #表を２次元配列として初期化しておく
    row_max = len(codelist)
    col_max = len(month_list)
    shouhi_hyo = [[''] * (col_max+1) for i in range(row_max+1)]

    #1行目は２列目以降にmonthヘッダを
    for i in range(col_max ):
        shouhi_hyo[0][i+1] = month_list[i][0]

    #1列目は２行目以降にcodelistのコードを
    for j in range(row_max ):
        shouhi_hyo[j+1][0] = codelist[j]

    #Shouhiデータから、該当する消費数量を記入する。
    shouhis = Shouhi.objects.all()
    for i in range(col_max):
        for j in range(row_max):
            for shouhi in shouhis:
                if shouhi_hyo[0][i+1] == shouhi.month and shouhi_hyo[j+1][0] == shouhi.code :
                    shouhi_hyo[j+1][i+1] = shouhi.qty

    #エクセルsheet2 に書き込み
    for i in range(col_max+1):
        for j in range(row_max+1):
            sheet2.cell(row=j+1, column=i+1, value = shouhi_hyo[j][i])

    save_file = 'TFC_kento_{0}.xlsx'.format(str(kijunbi).replace('/', '-'))
    return wb, save_file
        
#def write_kexcel(hyo, kijunbi, nolist):
#    filename = os.path.join(settings.MEDIA_ROOT, 'template', ZEXCEL)
#    wb = openpyxl.load_workbook(filename)
#    wb = write_kexcel_hyo(hyo, kijunbi, nolist, wb)
#    save_file = 'TFC_kento_{0}.xlsx'.format(str(kijunbi).replace('/', '-'))
#    return wb, save_file
        
#def write_kexcel_with_shouhi(hyo, kijunbi, nolist):
#    filename = os.path.join(settings.MEDIA_ROOT, 'template', ZEXCEL)
#    wb = openpyxl.load_workbook(filename)
#    wb = write_kexcel_hyo(hyo, kijunbi, nolist, wb)
#    wb = write_shouhi(wb)
#    save_file = 'TFC_kento_{0}.xlsx'.format(str(kijunbi).replace('/', '-'))
#    return wb, save_file

def write_zaiko(data, kijunbi, begin_day):

    k_data =  data #発注検討表データ取込み
    mb = MakeBalance(begin_day)
    nolist = mb.make_nolist()

    #在庫コードデータを取り出し。 get_zは在庫フラグ1のコードとカテゴリリスト
    c_data = ori_sort(get_z()) 
    zdata=join_data(c_data, k_data)
    #旧モデルで在庫と受注の無いものは省きます。
    zdata = [e for e in zdata if not (e[1] == '旧モデル' and e[2] == 0 and e[3] == 0)]
    #print('mb', mb.totallist)
    #print('nolist', nolist)
    hyo = make_hyo(nolist, zdata, mb.totallist)
    return write_zexcel(hyo, kijunbi, nolist)
        
    #with open('zdata.csv', 'w', encoding='CP932') as f:
    #    writer = csv.writer(f)
    #    writer.writerows(hyo)

def write_kento(data, kijunbi, begin_day):

    k_data =  data #発注検討表データ取込み
    mb = MakeBalance(begin_day)
    nolist = mb.make_nolist()

    c_data = ori_sort(get_k())
    kdata=join_data(c_data, k_data)
    hyo = make_hyo(nolist, kdata, mb.totallist)

    #with open('hyo.csv', 'w', encoding='CP932') as f:
    #    writer = csv.writer(f)
    #    writer.writerows(hyo)

    return write_kexcel(hyo, kijunbi, nolist)


