#po_linesファイルからデータを読みこんで、PO作成。
#発注データを保存する => データベースなど

import openpyxl
import os
from django.conf import settings
from .models import Poline, Fabric

POFILE = "po.xlsx"

ITEMROWBEGIN = 16 #アイテム行は１から数えて何番目から
ITEMCOLUMNBEGIN = 2 #アイテム列は、A=１から数えて何番目

PODATE ="J6"

MAX = 300 #ベースPOのアイテム行の最後行 index

def make_ouritem(hinban, fabs):
    for fab in fabs:
        hinban = hinban.replace(fab.code, fab.code + "[" + fab.name + "]")

    return hinban


#エクセルPOファイル雛形を読み込みデータを書き込む
def write_po_excel(po):
    filename = os.path.join(settings.MEDIA_ROOT, 'template', POFILE)
    book = openpyxl.load_workbook(filename)
    sheet = book['PO']
    r = ITEMROWBEGIN
    polines = Poline.objects.filter(po=po)
    fabs = Fabric.objects.all()
    for pl in polines:
        c = ITEMCOLUMNBEGIN
        sheet.cell(row=r, column=c).value = pl.code.item
        sheet.cell(row=r, column=c+1).value = pl.code.description
        sheet.cell(row=r, column=c+2).value = pl.remark
        sheet.cell(row=r, column=c+3).value = int(float(pl.qty))
        sheet.cell(row=r, column=c+4).value = pl.code.unit
        sheet.cell(row=r, column=c+5).value = (float(pl.code.uprice) 
            if pl.code.uprice.replace(".", "").isdigit() else 0)
        #金額計算式挿入
        sheet.cell(row=r, column=c+6).value = "=E{0}*G{0}".format(r)
        sheet.cell(row=r, column=c+8).value = make_ouritem(pl.code.hinban, fabs)
        #容積計算式挿入
        sheet.cell(row=r, column=c+9).value = "=E{0}*N{0}".format(r)
        sheet.cell(row=r, column=c+10).value = pl.om
        sheet.cell(row=r, column=c+12).value = (float(pl.code.vol) 
            if pl.code.vol.replace(".", "").isdigit() else 0)
        r += 1

    #書き込み最終 index から MAX 行-1行を削除
    c = ITEMCOLUMNBEGIN
    sheet.delete_rows(r+1, MAX-r )
    #数量合計式挿入
    sheet.cell(row=r+1, column=c+3).value = "=SUM(E{0}:E{1})".format(ITEMROWBEGIN, r) 
    #金額合計式挿入
    sheet.cell(row=r+1, column=c+6).value = "=SUM(H{0}:H{1})".format(ITEMROWBEGIN, r) 
    #容積合計式挿入
    sheet.cell(row=r+1, column=c+9).value = "=SUM(K{0}:K{1})".format(ITEMROWBEGIN, r) 
    sheet.merge_cells("E{0}:H{0}".format( r+3 ))

    #発注日挿入
    sheet.cell(row=6, column=10).value = po.pod

    #comment 挿入
    sheet.cell(row=r+2, column=3).value = po.comment

    #PONO挿入
    sheet.cell(row=7, column=10).value = po.pon

    #積載コメント挿入
    if po.ft40 is not None and len(str(po.ft40)) > 0 :
    #if po.ft40 is not None and po.ft40 > 0 :
        sheet.cell(row=r+1, column=3).value = "＊{}コンテナに積載してください。".format('40HC x ' + str(po.ft40))
    elif po.ft20 is not None and len(str(po.ft20)) > 0 :
    #elif po.ft20 is not None and po.ft20 > 0 :
        sheet.cell(row=r+1, column=3).value = "＊{}コンテナに積載してください。".format('20ft x ' + str(po.ft20))
    else:
        sheet.cell(row=r+1, column=3).value = po.condition.via

    #shipment per挿入
    sheet.cell(row=r+3, column=2).value = "Shipment per:"
    sheet.cell(row=r+3, column=3).value = po.condition.shipment_per

    #ship to挿入
    sheet.cell(row=r+4, column=2).value = "Ship to:"
    sheet.cell(row=r+4, column=3).value = po.condition.shipto_1

    #ship to 2 なしの場合
    if not po.condition.shipto_2 :
    #Via タイトル : データ挿入
        sheet.cell(row=r+5, column=2).value = "Via:"
        sheet.cell(row=r+5, column=3).value = po.condition.via
        #forwarder タイトル : データ挿入
        sheet.cell(row=r+6, column=2).value = "Forwarder: "
        sheet.cell(row=r+6, column=3).value = po.condition.forwarder
            
        #Delivery タイトル : データ挿入
        sheet.cell(row=r+3, column=4).value = "Delivery(ETD): "
        sheet.cell(row=r+3, column=5).value = po.etd
        #Tradeterm タイトル : データ挿入
        sheet.cell(row=r+4, column=4).value = "Trade Term: "
        sheet.cell(row=r+4, column=5).value = po.condition.trade_term
        #Payment タイトル : データ挿入
        sheet.cell(row=r+5, column=4).value = "Payment: "
        sheet.cell(row=r+5, column=5).value = po.condition.payment_term
        #Insurance タイトル : データ挿入
        sheet.cell(row=r+6, column=4).value = "Insurance: "
        sheet.cell(row=r+6, column=5).value = po.condition.insurance

        # サイン欄までの空白行を2行削除
        sheet.delete_rows(r+7, 2 )

    else:
        #ship to_2,3,4, 5挿入
        sheet.cell(row=r+5, column=3).value = po.condition.shipto_2
        sheet.cell(row=r+6, column=3).value = po.condition.shipto_3
        sheet.cell(row=r+7, column=3).value = po.condition.shipto_4
        sheet.cell(row=r+8, column=3).value = po.condition.shipto_5

        #Delivery タイトル : データ挿入
        sheet.cell(row=r+3, column=4).value = "Delivery(ETD): "
        sheet.cell(row=r+3, column=5).value = po.etd

        #Tradeterm タイトル : データ挿入
        sheet.cell(row=r+4, column=4).value = "Trade Term:: "
        sheet.cell(row=r+4, column=5).value = po.condition.trade_term

        #Payment タイトル : データ挿入
        sheet.cell(row=r+5, column=4).value = "Payment: "
        sheet.cell(row=r+5, column=5).value = po.condition.payment_term

        #Insurance タイトル : データ挿入
        sheet.cell(row=r+6, column=4).value = "Insurance: "
        sheet.cell(row=r+6, column=5).value = po.condition.insurance
            

    if po.ft40 is not None and len(str(po.ft40)) > 0 :
    #if po.ft40 is not None and po.ft40 > 0 :
        via = '40x' + str(po.ft40)
    elif po.ft20 is not None and len(str(po.ft20)) > 0 :
    #elif po.ft20 is not None and po.ft20 > 0 :
        via ='20x' + str(po.ft20)
    else:
        via = po.condition.shipment_per

    pofilename = "PO" + po.pon + po.etd.strftime("(%m%d_") + po.port.replace('Port', '').replace('Japanese', '') +'_' + via + '_' + po.condition.nic + ").xlsx"

    return book, pofilename

