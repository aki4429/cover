from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from .models import TfcCode, Juchu, Cart, Condition, Po, Poline, Inv, Invline, Kento, Shouhi
from loc.models import LocStatus
from .forms import CodeForm, JuchuForm, ConditionForm, PoForm, PolineForm, CartForm, InvUpForm, InvForm, InvlineForm, MakezaikoForm, KentoForm, SokoUpForm
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from .juchu_read_2 import JuchuRead
from .make_po import write_po_excel
import datetime, os
from django.http import HttpResponse
from bootstrap_datepicker_plus import DateTimePickerInput
import openpyxl
from .read_inv_3 import ReadInv
from .get_kh import read_kh
import datetime
from .write_zk import write_zaiko, write_kento
from .get_shouhi_new import read_shouhi
from .get_shouhi import read_shouhi_old

class CodeList(LoginRequiredMixin, ListView):
    context_object_name = 'codes'
    model = TfcCode

    #paginate_by = 10
    #検索語とフィールド名を指定して、その言葉がフィールドに
    #含まれるクエリーセットを返す関数。
    #検索語はスペースで区切って、リストでANDでつなげる
    def make_q(self,query, q_word, f_word):
        #(2)キーワードをリスト化させる(複数指定の場合に対応させるため)
        search      = self.request.GET[q_word].replace("　"," ")
        search_list = search.split(" ")
        filter = f_word + '__' + 'contains' #name__containsを作る
        #(例)info=members.filter(**{ filter: search_string })
        #(3)クエリを作る
        for word in search_list:
        #TIPS:AND検索の場合は&を、OR検索の場合は|を使用する。
            query &= Q(**{ filter: word })
            #(4)作ったクエリを返す
        return query


    def get_queryset(self):
        query = Q()
        flag = 0 #一つも検索語がなければ、flag==0
        if 'qh' in self.request.GET:
            query = self.make_q(query, 'qh', 'hinban')
            flag += 1

        if 'qi' in self.request.GET:
            query = self.make_q(query, 'qi', 'item')
            flag += 1

        if 'qs' in self.request.GET:
            query = self.make_q(query, 'qs', 'description')
            flag += 1

        if 'qb' in self.request.GET:
            query = self.make_q(query, 'qb', 'remarks')
            flag += 1

        if 'qc' in self.request.GET:
            query = self.make_q(query, 'qc', 'hcode')
            flag += 1

        if 'qt' in self.request.GET:
            query = self.make_q(query, 'qt', 'cat')
            flag += 1

        if 'z' in self.request.GET:
            query = self.make_q(query, 'z', 'zaiko')
            flag += 1

        if 'h' in self.request.GET:
            query = self.make_q(query, 'h', 'kento')
            flag += 1

        if flag == 0:
            codes = TfcCode.objects.order_by('hinban')
        else:
            print('Query',query)
            codes = TfcCode.objects.filter(query).order_by('hinban')

        return codes

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title']='TFCコードリスト'
        return context


class CodeCopy(LoginRequiredMixin, UpdateView):
    template_name = 'po/code_update.html'
    model = TfcCode
    form_class = CodeForm

    def get_success_url(self):
        return reverse('tfc_code')

    #ここでコピーするためにget_objectをオーバーライドします。
    def get_object(self, queryset=None):
        #self.request.GETは使えないので、self.kwargsを使うところがミソ
        tfccode = TfcCode.objects.get(pk=self.kwargs.get('pk'))
        #プライマリーキーを最新に。これがしたいためのオーバーライド
        tfccode.pk = TfcCode.objects.last().pk + 1
        return tfccode

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title']='TFCコードコピー'
        #context['last_id'] = TfcCode.objects.last().pk
        return context

class CodeUpdate(LoginRequiredMixin, UpdateView):
    template_name = 'po/code_update.html'
    model = TfcCode
    form_class = CodeForm

    def get_success_url(self):
        return reverse('tfc_code')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title']='TFCコード編集'
        #context['last_id'] = TfcCode.objects.last().pk
        return context

class PolineUpdate(LoginRequiredMixin, UpdateView):
    template_name = 'po/poline_update.html'
    model = Poline
    form_class = PolineForm

    def get_success_url(self):
        return reverse('poline_list', kwargs={'po_pk': self.object.po_id})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title']='PO内容編集'
        #context['last_id'] = TfcCode.objects.last().pk
        return context

class CartUpdate(LoginRequiredMixin, UpdateView):
    template_name = 'po/cart_update.html'
    model = Cart
    form_class = CartForm

    def get_success_url(self):
        return reverse('cart_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title']='カート編集'
        #context['last_id'] = TfcCode.objects.last().pk
        return context

class CodeDelete(LoginRequiredMixin, DeleteView):
    template_name = 'po/code_confirm_delete.html'
    model = TfcCode
    form_class = CodeForm

    def get_success_url(self):
        return reverse('tfc_code')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title']='TFCコード削除確認'
        return context

class PoDelete(LoginRequiredMixin, DeleteView):
    template_name = 'po/po_confirm_delete.html'
    model = Po
    form_class = PoForm

    def get_success_url(self):
        return reverse('po_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title']='PO削除確認'
        return context

class PolineDelete(LoginRequiredMixin, DeleteView):
    template_name = 'po/poline_confirm_delete.html'
    model = Poline
    #form_class = PolineForm

    def get_success_url(self):
        return reverse('poline_list', kwargs={'po_pk': self.object.po_id})
        #return reverse('po_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title']='PO行削除確認'
        return context

class CodeDetail(LoginRequiredMixin, DetailView):
    template_name = 'po/code_detail.html'
    model = TfcCode

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title']='TFCコード詳細'
        return context

class CodeCreate(LoginRequiredMixin, CreateView):
    template_name = 'po/code_create.html'
    model = TfcCode
    form_class = CodeForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title']='TFCコード作成'
        return context


class CartCreate(LoginRequiredMixin, CreateView):
    template_name = 'po/cart_create.html'
    model = Cart
    form_class = CartForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title']='TFCコード作成'
        return context

@login_required
def juchu_upload(request):
    if request.method == 'POST':
        form = JuchuForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('juchu_list')
    else:
        form = JuchuForm()

    context = {
        'title' : '受注データアップロード',
        'form': form
        }

    return render(request, 'po/juchu_upload.html', context )

@login_required
def soko_upload(request):
    if request.method == 'POST':
        form = SokoForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('soko_list')

    else:
        form = SokoForm()

    context = {
        'title' : '倉庫別在庫表アップロード',
        'form': form
        }

    return render(request, 'po/soko_upload.html', context )


class JuchuList(LoginRequiredMixin, ListView):
    context_object_name = 'juchus'
    template_name = 'po/juchu_list.html'
    model = Juchu
    form_class = JuchuForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title']='受注データリスト'
        return context

class ShouhiList(LoginRequiredMixin, ListView):
    context_object_name = 'shouhis'
    template_name = 'po/shouhi_list.html'
    model = Shouhi

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        ms = Shouhi.objects.all().order_by('month').values_list('month').distinct() 
        months = []
        for mt in ms:
            months.append(mt[0])

        context['title']='消費データリスト'
        context['months']= months
        return context

class PolineList_2(LoginRequiredMixin, ListView):
    context_object_name = 'polines'
    template_name = 'po/poline_list_2.html'
    model = Poline
    form_class = PolineForm

    def get_queryset(self):
        po = Po.objects.get(pk=self.kwargs.get('po_pk'))
        return Poline.objects.filter(po=po)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title']='PO明細リスト'
        return context

class KentoList(LoginRequiredMixin, ListView):
    context_object_name = 'kentos'
    template_name = 'po/kento_list.html'
    model = Kento
    form_class = KentoForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title']='発注検討表アップロードリスト'
        return context

@login_required
def juchu_delete(request, pk):
   juchu = get_object_or_404(Juchu, id=pk)
   juchu.delete()
   return redirect('juchu_list')

@login_required
def zkento_delete(request, pk):
   kento = get_object_or_404(Kento, id=pk)
   kento.delete()
   return redirect('kento_list')

#コードdetailからカートに追加
@login_required
def cart_append(request, pk):
   code = get_object_or_404(TfcCode, id=pk)
   cart = Cart(hinban=code.hinban, code=code)
   cart.save()
   return redirect('cart_update', pk = cart.pk)

@login_required
def make_cart(request, juchu_id):
    #Cart.objects.all().delete()
    juchu = get_object_or_404(Juchu, id=juchu_id)

    juchu_file_name =  juchu.file_name.path
    jr = JuchuRead(juchu_file_name)
    jdata = jr.get_juchu()

    add_cart = []
    for row in jdata:
        result = ''
        if type(row[6]) == int :
            result = TfcCode.objects.filter(pk=str(row[6]))
        elif type(row[6]) and row[6].isdigit():
            result = TfcCode.objects.filter(pk=str(row[6]))

        if len(result) == 1:
            code = result.first()
        else:
            code = None
        
        cart = Cart(hinban = row[0], \
                om = row[1],
                juchubi = row[2].replace('/','-'),
                noki = row[3].replace('/','-'),
                qty = int(float(row[4])),
                flag = row[5],
                code = code,
                obic = row[7])
        add_cart.append(cart)

    Cart.objects.bulk_create(add_cart)

    return redirect('cart_list')


class CartList(LoginRequiredMixin, ListView):
    context_object_name = 'cart'
    template_name = 'po/cart_list.html'
    model = Cart

    #paginate_by = 10
    #検索語とフィールド名を指定して、その言葉がフィールドに
    #含まれるクエリーセットを返す関数。
    #検索語はスペースで区切って、リストでANDでつなげる
    def make_q(self,query, q_word, f_word):
        #(2)キーワードをリスト化させる(複数指定の場合に対応させるため)
        search      = self.request.GET[q_word].replace("　"," ")
        search_list = search.split(" ")
        filter = f_word + '__' + 'contains' #name__containsを作る
        #(例)info=members.filter(**{ filter: search_string })
        #(3)クエリを作る
        for word in search_list:
        #TIPS:AND検索の場合は&を、OR検索の場合は|を使用する。
            query &= Q(**{ filter: word })
            #(4)作ったクエリを返す
        return query

    def get_queryset(self):
        query = Q()
        flag = 0 #一つも検索語がなければ、flag==0
        if 'qh' in self.request.GET:
            query = self.make_q(query, 'qh', 'hinban')
            flag += 1

        if 'qj' in self.request.GET:
            query = self.make_q(query, 'qj', 'om')
            flag += 1

        if 'qf' in self.request.GET:
            query = self.make_q(query, 'qf', 'flag')
            flag += 1

        if 'qo' in self.request.GET:
            query = self.make_q(query, 'qo', 'obic')
            flag += 1

        if flag == 0:
            cart = Cart.objects.order_by('om')
        else:
            print('Query',query)
            cart= Cart.objects.filter(query).order_by('om')

        return cart

    def post(self, request):
        orders = request.POST.getlist('order')  # <input type="checkbox" name="delete"のnameに対応
        #POSTされたorderの配列をセッションに保存
        request.session['orders'] = orders
        for cpk in orders:
            Cart.objects.filter(pk=cpk).update(flag='order') 

        if 'new_order' in request.POST:
            return redirect('condition_list')  # 一覧ページにリダイレクト
        elif 'add_order' in request.POST:
            return redirect('po_list')  # 一覧ページにリダイレクト

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title']='カートリスト'
        return context

class ConditionList(LoginRequiredMixin, ListView):
    context_object_name = 'conditions'
    template_name = 'po/condition_list.html'
    model = Condition
    form_class = ConditionForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title']='輸入条件選択'
        #ordersセッションが存在しないか、空のときorder_exitはFalse
        if 'orders' in self.request.session and len(self.request.session['orders']) > 0:
            context['order_exist']= True
        else:
            context['order_exist']= False

        return context

class ConditionUpdate(LoginRequiredMixin, UpdateView):
    template_name = 'po/condition_update.html'
    model = Condition
    form_class = ConditionForm

    def get_success_url(self):
        return reverse('condition_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title']='輸入条件編集'
        #context['last_id'] = TfcCode.objects.last().pk
        return context

class ConditionCopy(LoginRequiredMixin, UpdateView):
    template_name = 'po/condition_update.html'
    model = Condition
    form_class = ConditionForm

    def get_success_url(self):
        return reverse('condition_list')

    #ここでコピーするためにget_objectをオーバーライドします。
    def get_object(self, queryset=None):
        #self.request.GETは使えないので、self.kwargsを使うところがミソ
        condi = Condition.objects.get(pk=self.kwargs.get('pk'))
        #プライマリーキーを最新に。これがしたいためのオーバーライド
        condi.pk = Condition.objects.last().pk + 1
        return condi

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title']='輸入条件コピー'
        #context['last_id'] = TfcCode.objects.last().pk
        return context

class ConditionDelete(LoginRequiredMixin, DeleteView):
    template_name = 'po/condition_confirm_delete.html'
    model = Condition
    form_class = ConditionForm

    def get_success_url(self):
        return reverse('condition_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title']='輸入条件削除確認'
        return context

class CartDelete(LoginRequiredMixin, DeleteView):
    template_name = 'po/cart_confirm_delete.html'
    model = Cart

    def get_success_url(self):
        return reverse('cart_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title']='カート行削除確認'
        return context

@login_required
def cart_delete_all(request):
    Cart.objects.all().delete()
    return redirect('cart_list')

class PoCreate(LoginRequiredMixin, CreateView):
    template_name = 'po/po_create.html'
    model = Po
    form_class = PoForm

    def get_initial(self):
         initial = super().get_initial()

         #最新pon を取り出し、数字部分に1加えてHをつける
         new_pon = 'H' + str(int(Po.objects.last().pon.split('H')[1])+1)
         initial["pon"] = new_pon

         #発注日の初期値は今日の日付
         initial["pod"] = datetime.date.today

         #URLで受け取ったparameterから、Conditionのインスタンスを取り出し。
         condi = Condition.objects.get(pk=self.kwargs.get('condi_pk'))
         initial["condition"] = condi
         initial["port"] = condi.via
         initial["per"] = condi.shipment_per
         initial["shipto"] = condi.shipto_1
         initial["comment"] = condi.comment
         return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title']='PO作成'
        if 'orders' in self.request.session:
            context['orders']= self.request.session['orders']
        return context

    def get_success_url(self):
        #carts = Cart.objects.filter(flag='order')
        #orders = []
        #for cart in carts:
        #    orders.append(cart.pk)

        orders = self.request.session['orders']
        polines = []
        for cartpk in orders:
            cart = Cart.objects.get(pk=cartpk)
            #tfccode  = TfcCode.objects.get(pk=cart.code)
            tfccode  = cart.code
            pl = Poline(
                code = tfccode,
                remark =tfccode.remarks,
                om =cart.om,
                qty =cart.qty,
                balance =cart.qty,
                po = Po.objects.last()
                )
            polines.append(pl)

        Poline.objects.bulk_create(polines)

        return reverse('po_list')

class PoList(LoginRequiredMixin, ListView):
    context_object_name = 'pos'
    template_name = 'po/po_list.html'
    model = Po
    #form_class = PoForm

    def get_queryset(self):
        return Po.objects.filter(etd__gte='2020-01-01').order_by('-pon')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title']='POリスト'
        #add_ordersセッションが存在しないか、空のときorder_exitはFalse
        #if 'add_orders' in self.request.session and len(self.request.session['add_orders']) > 0:
        if 'orders' in self.request.session and len(self.request.session['orders']) > 0:
            context['add_order_exist']= True
        else:
            context['add_order_exist']= False

        return context

class PoUpdate(LoginRequiredMixin, UpdateView):
    context_object_name = 'pos'
    template_name = 'po/po_update.html'
    model = Po
    form_class = PoForm

    def get_success_url(self):
        return reverse('poline_list', kwargs={'po_pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title']='PO編集'
        return context

    #def get_form(self):
    #    form = super(PoUpdate, self).get_form()
    #    return form

class PolineList(LoginRequiredMixin, ListView):
    context_object_name = 'polines'
    template_name = 'po/poline_list.html'
    model = Poline
    #form_class = PoForm

    def get_queryset(self):
        po = Po.objects.get(pk=self.kwargs.get('po_pk'))
        return Poline.objects.filter(po=po)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title']='PO内容リスト'
        po = Po.objects.get(pk=self.kwargs.get('po_pk'))
        data=[]
        pls = Poline.objects.filter(po=po)
        pl_update =[] #bulk 一括update用にクラスを格納
        for pl in pls:
            ils = Invline.objects.filter(poline = pl)
            #引当済みのインボイス#と数量をinv_dataに追加代入
            inv_data=[]
            sum = 0
            for il in ils:
                inv_data.append(il.inv.invn)
                inv_data.append(il.qty)
                sum += il.qty
            line = []
            #残数を計算して更新
            pl.balance = pl.qty - sum
            pl_update.append(pl)
            #polineのpkは配列の[0]に代入
            line.append(pl.code.hinban)
            line.append(pl.remark)
            line.append(pl.qty)
            line.append(pl.om)
            line.append(pl.balance)
            #インボイスデータを追加
            line.extend(inv_data)

            data.append(line)

        Poline.objects.bulk_update(pl_update, fields=['balance'])
        context['po']= po
        context['data']= data
        return context

#POを作ってダウンロードします。
@login_required
def make_po(request, po_pk):
    po = get_object_or_404(Po, pk=po_pk)
    wb, file_name = write_po_excel(po)
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename={}'.format(file_name)
    wb.save(response)
    return response

@login_required
def order_list(request):
    if request.session['orders'] :
        orders = request['orders']
        context['orders']= orders
        return render(request, 'po/order_list.html', context)

#検討表をアップロードします。
@login_required
def kento_upload(request):
    context={}
    if request.method == 'POST' and request.FILES['excel']:
        excel = request.FILES['excel']
        #excelの読み込み
        wb = openpyxl.load_workbook(excel)
        sheet = wb['kento']

        #max行取得
        max = sheet.max_row

        #行列番号は1始まり、Row 5から、Col 2=コード、25=(Y列)今回発注
        #発注辞書に保存
        add_cart = []

        i=0
        while((i+5) <= max):
            code = sheet.cell(row=i+5, column=2).value
            #qty = sheet.cell(row=i+5, column=26).value  #Z
            #qty = sheet.cell(row=i+5, column=25).value  #Y
            #qty = sheet.cell(row=i+5, column=27).value  #AA
            #qty = sheet.cell(row=i+5, column=28).value  #AB
            #qty = sheet.cell(row=i+5, column=29).value  #AC
            qty = sheet.cell(row=i+5, column=30).value  #AD
            #qty = sheet.cell(row=i+5, column=31).value  #AE
            #print(type(qty), qty, not qty)
            if qty :
                tc = TfcCode.objects.filter(hcode__iexact=code)[0]
                cart = Cart(hinban = tc.hinban, \
                    qty = int(float(qty)),
                    code = tc
                    )
                add_cart.append(cart)
            i += 1

        Cart.objects.bulk_create(add_cart)

        context = {
            'title' : '発注検討表アップロード',
            }
        #os.remove(excel)

    return render(request, 'po/kento_upload.html', context )

@login_required
def add_order(request, po_pk):
    po = get_object_or_404(Po, pk=po_pk)
    #orders = request.session['add_orders']
    orders = request.session['orders']
    polines = []
    for cartpk in orders:
        cart = Cart.objects.get(pk=cartpk)
        tfccode  = cart.code
        pl = Poline(
        code = tfccode,
        remark =tfccode.remarks,
        om =cart.om,
        qty =cart.qty,
        balance =cart.qty,
        po = po
        )

        polines.append(pl)

    Poline.objects.bulk_create(polines)
    return redirect('poline_list', po_pk = po.pk)

@login_required
def upload_inv(request):
    INVN = (2,0) #inv.#位置
    form = InvUpForm()

    if request.method == 'POST':
        form = InvUpForm(request.POST, request.FILES)  # Do not forget to add: request.FILES
        if form.is_valid():
            file, download_url = form.save()
            r = ReadInv(file.name)
            context = {
                'invn': r.invn,
                'etd': r.etd,
                'form': form,
                }
            os.remove(file.name)
            return redirect('inv_list')
    return render(request, 'po/upload_inv.html', locals())

#倉庫別在庫表をアップロードして消費実績DB登録
@login_required
def upload_soko(request):
    form = SokoUpForm()

    if request.method == 'POST':
        form = SokoUpForm(request.POST, request.FILES)  # Do not forget to add: request.FILES
        if form.is_valid():
            file, download_url = form.save()
            read_shouhi(file.name)
            os.remove(file.name)
            return redirect('shouhi_list')
    return render(request, 'po/upload_soko.html', locals())

#旧倉庫別在庫表をアップロードして消費実績DB登録
@login_required
def upload_soko_old(request):
    form = SokoUpForm()

    if request.method == 'POST':
        form = SokoUpForm(request.POST, request.FILES)  # Do not forget to add: request.FILES
        if form.is_valid():
            file, download_url = form.save()
            read_shouhi_old(file.name)
            os.remove(file.name)
            return redirect('shouhi_list')
    return render(request, 'po/upload_soko_old.html', locals())

class InvList(LoginRequiredMixin, ListView):
    context_object_name = 'invs'
    template_name = 'po/inv_list.html'
    model = Inv
    #form_class = PoForm

    def get_queryset(self):
        return Inv.objects.filter(etd__gte='2020-01-01').order_by('-etd')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title']='INVリスト'
        return context

class InvlineList(LoginRequiredMixin, ListView):
    context_object_name = 'invlines'
    template_name = 'po/invline_list.html'
    model = Invline

    def get_queryset(self):
        inv = Inv.objects.get(pk=self.kwargs.get('inv_pk'))
        return Invline.objects.filter(inv=inv)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title']='INV内容リスト'
        inv = Inv.objects.get(pk=self.kwargs.get('inv_pk'))
        context['inv']= inv
        return context

class InvUpdate(LoginRequiredMixin, UpdateView):
    template_name = 'po/inv_update.html'
    model = Inv
    form_class = InvForm

    #fields = ['invn', 'etd', 'delivery']
 
    """
    def get_form(self):
        form = super(InvUpdate, self).get_form()
        form.fields['invn'].label = 'インボイスNo.'
        form.fields['etd'].label = 'ETD'
        form.fields['delivery'].label = '取込日'
        return form
   """

    def get_success_url(self):
        return reverse('invline_list', kwargs={'inv_pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title']='INV編集'
        context['inv_form'] = InvForm
        return context

class InvDelete(LoginRequiredMixin, DeleteView):
    template_name = 'po/inv_confirm_delete.html'
    model = Inv
    form_class = InvForm

    def get_success_url(self):
        return reverse('inv_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title']='INV削除確認'
        return context

class InvineUpdate(LoginRequiredMixin, UpdateView):
    template_name = 'po/invline_update.html'
    model = Invline
    form_class = InvlineForm

    def get_success_url(self):
        return reverse('invline_list', kwargs={'inv_pk': self.object.inv.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title']='Inv内容編集'
        return context

class InvlineDelete(LoginRequiredMixin, DeleteView):
    template_name = 'po/invline_confirm_delete.html'
    model = Invline
    #form_class = PolineForm

    def get_success_url(self):
        return reverse('invline_list', kwargs={'inv_pk': self.object.inv_id})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title']='INV行削除確認'
        return context

class InvlineUpdate(LoginRequiredMixin, UpdateView):
    template_name = 'po/invline_update.html'
    model = Invline
    form_class = InvlineForm

    def get_success_url(self):
        return reverse('invline_list', kwargs={'inv_pk': self.object.inv_id})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title']='INV内容編集'
        return context


@login_required
def make_zaiko(request, kento_id):
    params = {'begin_date': '', 'kijunbi': '', 'title':'在庫表作成', 'form': None}
    kento = get_object_or_404(Kento, id=kento_id)

    kento_file_name =  kento.file_name.path
    data, kijunbi = read_kh(kento_file_name)
    kijunbi = datetime.datetime.strptime(kijunbi, '%Y/%m/%d').date()

    status = LocStatus.objects.get(pk=1)
    shijibi = status.shijibi
    shijibi = datetime.datetime.strptime(shijibi, '%Y%m%d')
    koshinbi = status.koshinbi

    if request.method == 'POST':
        form = MakezaikoForm(request.POST)
        begin_day = request.POST['begin_date']
        params['begin_date'] = begin_day
        params['data'] = data
        params['shijibi'] = shijibi
        params['koshinbi'] = koshinbi
        params['kijunbi'] = kijunbi
        params['form'] = form
        params['message'] = 'POSTされました！'
        print('begin_day', begin_day)
        if 'zaiko' in request.POST:
            wb, file_name = write_zaiko(data, kijunbi, begin_day)
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename={}'.format(file_name)
            wb.save(response)
            return response
        elif 'kento' in request.POST:
            wb, file_name = write_kento(data, kijunbi, begin_day)
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename={}'.format(file_name)
            wb.save(response)
            return response
    else:
        params['data'] = data
        params['shijibi'] = shijibi
        params['koshinbi'] = koshinbi
        params['kijunbi'] = kijunbi
        params['form'] = MakezaikoForm()
    return render(request, 'po/make_zaiko.html', params)

@login_required
def zkento_upload(request):
    if request.method == 'POST':
        form = KentoForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('kento_list')
    else:
        form = KentoForm()

    context = {
        'title' : '在庫表用検討表データアップロード',
        'form': form
        }

    return render(request, 'po/zkento_upload.html', context )

