import openpyxl, os
from cover.settings import *

PEXCEL = 'pick_template.xlsx'

#受け取るpicks = Pick.objects.order_by('banch', 'code')
def write_pick_excel(picks):
    filename = os.path.join(MEDIA_ROOT, 'template', PEXCEL)
    wb = openpyxl.load_workbook(filename)

    sheet = wb['pick']
    #sheet['A2'] = "[ピッキング番地明細]" 

    #番地、コード、ピック数量、番地在庫、生産日、受注NO

    n = 2 #2行目からスタート
    for case in picks:
        sheet.cell(row=n, column=1, value = case.banch) 
        sheet.cell(row=n, column=2, value = case.code) 
        sheet.cell(row=n, column=3, value = case.qty) 
        sheet.cell(row=n, column=4, value = case.loc_qty) 
        sheet.cell(row=n, column=5, value = case.seisan) 
        sheet.cell(row=n, column=6, value = case.om) 
        n += 1

    return wb
