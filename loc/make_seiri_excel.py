import openpyxl, os
from cover.settings import *

SEXCEL = 'arrive_template.xlsx'

#受け取るinputs = Input.objects.order_by('hcode')
def write_excel(inputs, invn, bangos):
    filename = os.path.join(MEDIA_ROOT, 'template', SEXCEL)
    wb = openpyxl.load_workbook(filename)

    sheet = wb['yotei_out']
    sheet['A2'] = "入荷カバーリスト / " + invn

    #コード、数量、番地、数量記入
    n = 0 #5行目からスタート
    for case in inputs:
        sheet.cell(row=n+5, column=1, value = case.hcode) #code
        sheet.cell(row=n+5, column=2, value = case.qty) #qty
        sheet.cell(row=n+5, column=3, value = case.banch) #banch
        sheet.cell(row=n+5, column=4, value = case.kqty) #qty
        if int(float(case.kqty)) == 0:
            sheet.cell(row=n+5, column=5, value = '新規') #新規/既存
        else:
            sheet.cell(row=n+5, column=5, value = '既存') #新規/既存

        for ban in bangos:
            if ban.hcode == case.hcode.replace('013CH', '013').replace('232WI', '232W').replace('271I', '271'):
                sheet.cell(row=n+5, column=6, value = ban.se) #背番号

        n += 1
    return wb
